import os
import streamlit as st
import pandas as pd
import pg8000
from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Function to generate coversheets and save them to a zip file
def generate_coversheets_zip(student_list=[]):
    db_connection = pg8000.connect(
        database=os.environ["SUPABASE_DB_NAME"],
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        host=os.environ["SUPABASE_HOST"],
        port=os.environ["SUPABASE_PORT"]
    )

    db_cursor = db_connection.cursor()
    student_list_string = ', '.join(map(str, student_list))

    db_query = f"""SELECT student_list.name,                  
                    student_list.iatc_id,
                    student_list.nat_id,
                    student_list.class,
                    exam_list.exam_long AS subject,
                    exam_results.score,
                    exam_results.result,
                    exam_results.date
                    FROM exam_results 
                    JOIN student_list ON exam_results.nat_id = student_list.nat_id
                    JOIN exam_list ON exam_results.exam = exam_list.exam
                    WHERE student_list.iatc_id IN ({student_list_string}) AND exam_results.score_index = 1
                    ORDER BY exam_list.srt_exam ASC
                """
    db_cursor.execute(db_query)
    output_data = db_cursor.fetchall()
    db_cursor.close()
    db_connection.close()

    col_names = ['Name', 'IATC ID', 'National ID', 'Class', 'Subject', 'Score', 'Result', 'Date']
    df = pd.DataFrame(output_data, columns=col_names)

    # Create an in-memory ZIP file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for student_id in student_list:
            filtered_df = df[df['IATC ID'] == student_id]

            # Create a workbook and sheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = str(student_id)

            # Populate static text in specific cells
            sheet["B2"] = "Student Name:"
            sheet["B3"] = "Student IATC ID:"
            sheet["B4"] = "Student National ID:"
            sheet["B5"] = "Student Class:"

            # Populate specific values in the corresponding cells
            sheet["C2"] = filtered_df['Name'].iloc[0]
            sheet["C3"] = filtered_df['IATC ID'].iloc[0]
            sheet["C4"] = filtered_df['National ID'].iloc[0]
            sheet["C5"] = filtered_df['Class'].iloc[0]

            # Populate the data table starting from B7
            for col_num, header in enumerate(['Subject', 'Score', 'Result', 'Date'], start=2):
                sheet.cell(row=6, column=col_num, value=header).font = Font(bold=True)
                sheet.cell(row=6, column=col_num).alignment = Alignment(horizontal="center")

            for row_num, row_data in enumerate(filtered_df[['Subject', 'Score', 'Result', 'Date']].values, start=7):
                for col_num, value in enumerate(row_data, start=2):
                    sheet.cell(row=row_num, column=col_num, value=value)

            # Adjust column widths
            for column in ["B", "C", "D", "E"]:
                sheet.column_dimensions[column].width = 20

            # Save the workbook to a buffer
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)

            # Save Excel file in the zip
            excel_filename = f"{student_id}.xlsx"
            excel_buffer.seek(0)
            zip_file.writestr(excel_filename, excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer

# Streamlit interface
st.title("Generate Theory Exam Coversheets")
st.write("Enter a list of student IDs and download the Excel coversheets containing the highest result for each subject the student has taken.")

student_ids_input = st.text_area("Enter Student IDs separated by commas (e.g., 151596, 156756, 154960):")
st.write("Need help generating a list of IDs? Download the Excel template:")

# Direct link to the Excel file in GitHub
template_url = "https://github.com/Bayr-Harrison/coversheetgenerator/raw/main/Coversheet%20Generator%20Input.xlsx"
st.markdown(f"[Download Excel Template]({template_url})", unsafe_allow_html=True)

if st.button("Generate Coversheets"):
    try:
        student_list = [int(id.strip()) for id in student_ids_input.split(",")]
        st.write("Generating coversheets...")

        # Generate the zip file in memory
        zip_file = generate_coversheets_zip(student_list)

        # Offer the zip file for download
        st.download_button(
            label="Download All Coversheets as ZIP",
            data=zip_file,
            file_name="coversheets.zip",
            mime="application/zip"
        )
        st.success("Coversheets zip generated successfully!")

    except Exception as e:
        st.error(f"An error occurred: {e}")
